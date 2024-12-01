import requests
import json
import openpyxl

class LLM_preprocessing():
    def __init__(self, xlsx_filename: str):
        try:
            book = openpyxl.load_workbook(xlsx_filename)
            self.sheet = book.active
            self.start_processing()
        except FileNotFoundError:
            print("Ошибка загрузки файла xlsx")

    def start_processing(self) -> None:
        last_row_idx = self.sheet.max_row
        for idx in range(1, last_row_idx - 1):
            self.header = self.sheet[f'A{idx}'].value
            self.header = self.header.replace('\n', ' ')
            decline_reasons = self.sheet[f'B{idx}'].value
            self.llm_handle(decline_reasons)
            self.save_xlsx()
    
    def llm_handle(self, decline_reasons: str) -> str:
        numered_text = decline_reasons.split('\n')
        for idx in range(len(numered_text)):
            numered_text[idx] = str(idx) + ': ' + numered_text[idx]
        numered_text = '\n'.join(numered_text)
        prompt = f"""Ты исследователь текстов, который абсолютно точно соблюдает инструкции.
Необходимо проанализировать текст:\n{numered_text}\n
Каждая строка в тексте начинается со своего номера в формате '1: '.
В тексте нужно определить:
- номер строки, с которой начинается перечень причин отказа в приеме документов. 
- номер строки, на которой этот перечень причин отказа в приеме документов заканчивается
- номер строки, с которой начинается перечень причин отказа в предоставлении услуги
- номер строки, на которой этот перечень причин отказа в предоставлении услуги заканчивается
В ответе вывести два номера через запятую, без каких-либо дополнительных комментариев.
Никаких дополнительных комментариев не нужно! Только цифры через запятую.
Примеры ответа:
12, 33, 35, 45"""
        data = {
            "auth_token": "cupcoders_ze_best",
            "prompt": prompt
        }
        json_data = json.dumps(data, ensure_ascii=False).encode('utf8')
        r = requests.post(
            'http://25.35.193.226:5000/answer/', data=json_data)
        if r.status_code != 200:
            self.decline_reasons = ('', '')
            return
        answer = (r.content.decode())[1:-1]
        numbers_list = list(map(int, answer.split(", ")))
        splitted_text = decline_reasons.split('\n')
        decline_reasons = (
            ' '.join(splitted_text[numbers_list[0]:numbers_list[2]]),
            ' '.join(splitted_text[numbers_list[2]:-1])
        )
        self.decline_reasons = decline_reasons 
    
    def save_xlsx(self) -> None:
        try:
            book = openpyxl.load_workbook("final.xlsx")
            sheet = book.active
        except FileNotFoundError:
            book = openpyxl.Workbook()
            sheet = book.active
        last_row_idx = sheet.max_row
        sheet.merge_cells(f'A{last_row_idx+1}:B{last_row_idx+1}')
        sheet[f'A{last_row_idx+1}'] = str(self.header)
        sheet[f'A{last_row_idx+2}'] = "Отказ приема документов"
        sheet[f'B{last_row_idx+2}'] = "Отказ предоставления услуги"
        sheet[f'A{last_row_idx+3}'] = str(self.decline_reasons[0])
        sheet[f'B{last_row_idx+3}'] = str(self.decline_reasons[1])
        book.save("final.xlsx")

