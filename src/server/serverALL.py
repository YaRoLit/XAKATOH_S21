import re
import PyPDF2
import openpyxl
from flask import Flask, request, send_file, jsonify
from flask_cors import CORS
import os
from io import BytesIO

app = Flask(__name__)
CORS(app)

UPLOAD_FOLDER = 'uploads'
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

class PDF_parser():
    def __init__(self, pdf_filename: str, sheet) -> None:
        self.pdf_filename = pdf_filename
        self.sheet = sheet
        self.extract_text_from_pdf(0)
        self.service_name = self.extract_service_name()
        self.extract_text_from_pdf(10)
        self.text = self.find_application_two()
        self.insert_to_xlsx()

    def extract_text_from_pdf(self, start_page=10) -> None:
        with open(self.pdf_filename, 'rb') as file:
            reader = PyPDF2.PdfReader(file)
            text = ""
            for page_num in range(start_page - 1, len(reader.pages)):
                page = reader.pages[page_num]
                text += page.extract_text()
        self.text = text

    def extract_service_name(self) -> str:
        pattern = r"(«([^«»]|\n)*»)"
        match = re.search(pattern, self.text)
        if match:
            return match.group()
        return None

    def find_application_two(self):
        pattern = r"п\s*р\s*и\s*л\s*о\s*ж\s*е\s*н\s*и\s*е\s*№?\s*2"
        match = re.search(pattern, self.text, re.IGNORECASE)
        if match:
            start_index = match.start()
            pattern_end = r"п\s*р\s*и\s*л\s*о\s*ж\s*е\s*н\s*и\s*е\s*№?\s*3"
            match_end = re.search(pattern_end, self.text[start_index:], re.IGNORECASE)
            if match_end:
                end_index = match_end.start() + start_index
            else:
                end_index = len(self.text)
            return self.text[start_index:end_index]
        return None

    def insert_to_xlsx(self) -> None:
        last_row_idx = self.sheet.max_row + 1
        self.sheet[f'A{last_row_idx}'] = str(self.service_name)
        self.sheet[f'B{last_row_idx}'] = str(self.text)

@app.route('/upload', methods=['POST'])
def upload_files():
    if 'file' not in request.files:
        return jsonify({'error': 'No file found in the request'}), 400
    
    files = request.files.getlist('file')  # Получаем список файлов

    # Проверяем, существует ли уже файл output.xlsx
    xlsx_filename = os.path.join(UPLOAD_FOLDER, 'output.xlsx')
    
    # Если файл существует, открываем его для добавления данных, если нет — создаем новый
    if os.path.exists(xlsx_filename):
        workbook = openpyxl.load_workbook(xlsx_filename)
        sheet = workbook.active
    else:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Processed Data"
        # Добавляем заголовки в новый файл
        sheet['A1'] = 'Service Name'
        sheet['B1'] = 'Application Text'

    # Обрабатываем каждый файл
    for file in files:
        # Сохраняем каждый PDF файл
        pdf_path = os.path.join(UPLOAD_FOLDER, file.filename)
        file.save(pdf_path)

        # Создаем объект PDF_parser и вставляем данные в Excel
        try:
            PDF_parser(pdf_filename=pdf_path, sheet=sheet)
        except Exception as e: print(e)
            

    # Сохраняем Excel файл
    workbook.save(xlsx_filename)

    # Отправляем файл обратно пользователю
    return send_file(xlsx_filename, as_attachment=True, download_name="output.xlsx")

@app.route('/export-db', methods=['GET'])
def export_db():
    # Путь к файлу output.xlsx
    xlsx_filename = os.path.join(UPLOAD_FOLDER, 'output.xlsx')

    # Проверяем, существует ли файл
    if not os.path.exists(xlsx_filename):
        return jsonify({'error': 'Файл не найден.'}), 404

    # Отправляем файл пользователю
    return send_file(xlsx_filename, as_attachment=True, download_name="output.xlsx")

if __name__ == '__main__':
    app.run(debug=True)
